#!/usr/bin/env python3
"""
Chem-Dry Legal Compliance Checker
----------------------------------
Reads the violation report xlsx, fetches each flagged page URL,
and checks whether the original violation sentence is still present.

Usage:
    python compliance_checker.py <path_to_xlsx> [--output results.xlsx] [--delay 1.5]

Output:
    An Excel report with a STATUS column:
      FIXED       - violation sentence no longer found on the page
      STILL LIVE  - violation sentence still detected
      PAGE ERROR  - could not fetch the page (4xx / 5xx / timeout)
      SKIPPED     - duplicate URL+sentence combo already checked
"""

import argparse
import sys
import time
import unicodedata
import re


import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font

# ── Colours for the output sheet ──────────────────────────────────────────────
FILL_FIXED      = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_STILL_LIVE = PatternFill("solid", fgColor="FFC7CE")   # red
FILL_ERROR      = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_SKIPPED    = PatternFill("solid", fgColor="D9D9D9")   # grey

FONT_BOLD = Font(bold=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    """Collapse whitespace and normalise Unicode so HTML text matches xlsx text."""
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def extract_visible_text(html: str) -> str:
    """Return all visible text from an HTML page as a single normalised string."""
    soup = BeautifulSoup(html, "html.parser")
    # Remove script / style / meta noise
    for tag in soup(["script", "style", "noscript", "meta", "head"]):
        tag.decompose()
    return normalize(soup.get_text(separator=" "))


def fetch_page(url: str, timeout: int = 15) -> tuple[str | None, str]:
    """
    Fetch a URL and return (visible_text, status_message).
    Returns (None, error_msg) on failure.
    """
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code >= 400:
            return None, f"HTTP {r.status_code}"
        return extract_visible_text(r.text), "OK"
    except requests.exceptions.Timeout:
        return None, "Timeout"
    except requests.exceptions.SSLError as e:
        return None, f"SSL error: {e}"
    except requests.exceptions.ConnectionError as e:
        return None, f"Connection error: {e}"
    except Exception as e:
        return None, str(e)


def sentence_still_present(page_text: str, violation_sentence: str) -> bool:
    """
    Check whether a meaningful fragment of the violation sentence
    still appears in the page text.

    Strategy: take the first 120 characters of the normalised sentence
    as a fingerprint — long enough to be specific, short enough to
    survive minor rewording at the tail end.
    """
    norm_violation = normalize(violation_sentence)
    if not norm_violation:
        return False
    fingerprint = norm_violation[:120]
    return fingerprint in page_text


# ── Main ───────────────────────────────────────────────────────────────────────

def load_violations(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("ERROR: spreadsheet appears empty.")

    # Auto-detect header row (first row that contains 'Page URL' somewhere)
    header_row_idx = None
    for i, row in enumerate(rows):
        if any("page url" in str(c).lower() for c in row if c):
            header_row_idx = i
            break
    if header_row_idx is None:
        sys.exit("ERROR: Could not find a header row containing 'Page URL'. Check your column names.")

    headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]

    def col(name_fragment: str) -> int:
        """Return 0-based column index for first header containing name_fragment."""
        nf = name_fragment.lower()
        for idx, h in enumerate(headers):
            if nf in h.lower():
                return idx
        raise ValueError(f"Could not find column matching '{name_fragment}' in headers: {headers}")

    try:
        c_dba       = col("franchise dba")
        c_main      = col("franchise main")
        c_domain    = col("domain")
        c_page_url  = col("page url")
        c_keyword   = col("keyword")
        c_sentence  = col("violation content")
    except ValueError as e:
        sys.exit(f"ERROR: {e}")
violations = []
for row in rows[header_row_idx + 1:]:
    if not any(row):
        continue
    try:
    violations.append({
        "franchise_dba":  str(row[c_dba]   or "").strip(),
        "franchise_main": str(row[c_main]  or "").strip(),
        "domain":         str(row[c_domain] or "").strip(),
        "page_url":       str(row[c_page_url] or "").strip(),
        "keyword":        str(row[c_keyword] or "").strip(),
        "sentence":       str(row[c_sentence] or "").strip(),
    })
except IndexError:
    print(f"  WARNING: skipping malformed row (too few columns): {row}")

    wb.close()
    return violations


def run_check(violations: list[dict], delay: float) -> list[dict]:
    """Fetch each unique page URL once, then evaluate all violations against it."""
    # Cache page text by URL so we only fetch each page once
    page_cache: dict[str, tuple[str | None, str]] = {}
    results = []
    seen_pairs: set[tuple[str, str]] = set()

    total = len(violations)
    for i, v in enumerate(violations, 1):
        url      = v["page_url"]
        sentence = v["sentence"]
        pair     = (url, normalize(sentence)[:120])

        print(f"[{i}/{total}] {url}")

        # Skip true duplicates (same URL + same sentence fingerprint)
        if pair in seen_pairs:
            v["status"]  = "SKIPPED"
            v["detail"]  = "Duplicate row"
            results.append(v)
            continue
        seen_pairs.add(pair)

        # Fetch page (from cache if already retrieved)
        if url not in page_cache:
            if page_cache:           # don't sleep before very first request
                time.sleep(delay)
            page_text, fetch_status = fetch_page(url)
            page_cache[url] = (page_text, fetch_status)
            print(f"         → fetched ({fetch_status})")
        else:
            page_text, fetch_status = page_cache[url]
            print(f"         → cached  ({fetch_status})")

        if page_text is None:
            v["status"] = "PAGE ERROR"
            v["detail"] = fetch_status
        elif sentence_still_present(page_text, sentence):
            v["status"] = "STILL LIVE"
            v["detail"] = "Violation sentence detected on page"
        else:
            v["status"] = "FIXED"
            v["detail"] = "Violation sentence not found"

        results.append(v)

    return results


def write_report(results: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Results"

    col_headers = [
        "Franchise DBA", "Franchise Main", "Domain",
        "Page URL With Violation", "Red Flag Keyword",
        "Violation Content / Sentence", "STATUS", "Detail"
    ]
    ws.append(col_headers)
    for cell in ws[1]:
        cell.fill  = PatternFill("solid", fgColor="1F4E79")
        cell.font  = Font(bold=True, color="FFFFFF")

    STATUS_FILL = {
        "FIXED":      FILL_FIXED,
        "STILL LIVE": FILL_STILL_LIVE,
        "PAGE ERROR":  FILL_ERROR,
        "SKIPPED":    FILL_SKIPPED,
    }

    summary = {"FIXED": 0, "STILL LIVE": 0, "PAGE ERROR": 0, "SKIPPED": 0}

    for r in results:
        status = r.get("status", "")
        row_data = [
            r["franchise_dba"],
            r["franchise_main"],
            r["domain"],
            r["page_url"],
            r["keyword"],
            r["sentence"],
            status,
            r.get("detail", ""),
        ]
        ws.append(row_data)
        fill = STATUS_FILL.get(status, PatternFill())
        for cell in ws[ws.max_row]:
            cell.fill = fill
        summary[status] = summary.get(status, 0) + 1

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 80)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Status", "Count"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for status, count in summary.items():
        ws2.append([status, count])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"\nReport saved → {output_path}")
    print("Summary:")
    for k, v in summary.items():
        print(f"  {k:<12} {v}")


def main():
    parser = argparse.ArgumentParser(description="Chem-Dry Legal Compliance Checker")
    parser.add_argument("xlsx",             help="Path to the violation report (.xlsx)")
    parser.add_argument("--output", "-o",   default="compliance_results.xlsx",
                        help="Output file path (default: compliance_results.xlsx)")
    parser.add_argument("--delay",  "-d",   type=float, default=1.5,
                        help="Seconds to wait between page fetches (default: 1.5)")
    args = parser.parse_args()

    print(f"Loading violations from: {args.xlsx}")
    violations = load_violations(args.xlsx)
    print(f"  Found {len(violations)} violation rows across "
          f"{len(set(v['page_url'] for v in violations))} unique URLs\n")

    results = run_check(violations, delay=args.delay)
    write_report(results, args.output)


if __name__ == "__main__":
    main()
